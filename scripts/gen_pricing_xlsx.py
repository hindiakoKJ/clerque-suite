"""Build docs/Clerque-Pricing-Tiers.xlsx from canonical PLAN_CAPS (plan_caps.tmp.json).
Two sheets: a Solo-vs-Solo-Books comparison (customer-facing) + a full plan reference.
Styling kept light/clean — the user applies final graphics."""
import json, os, subprocess
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))

# Pull canonical plan data straight from the built shared-types PLAN_CAPS so the
# sheet can never drift from what the system actually bills.
_NODE = r'''
const { PLAN_CAPS, planLabel } = require("./packages/shared-types/dist/plans.js");
const out = {};
for (const code of Object.keys(PLAN_CAPS)) {
  const c = PLAN_CAPS[code];
  out[code] = { label: planLabel(code), monthly: Math.round(c.pricePhpMonthlyCents/100),
    moduleCount: c.moduleCount, baseSeats: c.baseSeats, maxTotal: c.maxTotal,
    annualMonths: c.annualMonthEquivalent };
}
process.stdout.write(JSON.stringify(out));
'''
caps = json.loads(subprocess.check_output(["node", "-e", _NODE], cwd=ROOT, text=True))

BROWN = "8B5E3C"; BROWN_DK = "5C3D26"; CREAM = "EEE9DF"; CREAM_LT = "F7F4EE"
GREEN = "2E7D32"; GREY = "9CA3AF"; WHITE = "FFFFFF"
ARIAL = "Arial"
thin = Side(style="thin", color="D9CFC0")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
center = Alignment(horizontal="center", vertical="center", wrap_text=True)
left = Alignment(horizontal="left", vertical="center", wrap_text=True)

def f(sz=10, b=False, color="000000"):
    return Font(name=ARIAL, size=sz, bold=b, color=color)

wb = Workbook()

# ─────────────────────────── Sheet 1: Pricing (comparison) ───────────────────
ws = wb.active
ws.title = "Pricing"
ws.sheet_view.showGridLines = False
ws.column_dimensions["A"].width = 42
ws.column_dimensions["B"].width = 22
ws.column_dimensions["C"].width = 22

solo = caps["SOLO_PRO"]; books = caps["SOLO_BOOKS"]

# Title
ws.merge_cells("A1:C1")
ws["A1"] = "Clerque Counter — Pricing"
ws["A1"].font = f(16, True, WHITE); ws["A1"].fill = PatternFill("solid", fgColor=BROWN)
ws["A1"].alignment = center; ws.row_dimensions[1].height = 30

YES, NO = "✓", "—"  # ✓  —

rows = [
    ("header", "Plan", solo["label"], books["label"]),
    ("price",  "Price / month", solo["monthly"], books["monthly"]),
    ("text",   "Best for", "The complete point of sale", "Full POS + simple bookkeeping"),
    ("section","POINT OF SALE", "", ""),
    ("feat",   "Users / cashiers", str(solo["maxTotal"]), str(books["maxTotal"])),
    ("feat",   "Unlimited products, recipes & FEFO inventory", YES, YES),
    ("feat",   "GCash · Maya · QR Ph · card tendering", YES, YES),
    ("feat",   "BIR-compliant Z-read & receipts", YES, YES),
    ("feat",   "PWD / Senior discounts", YES, YES),
    ("feat",   "Audit log, custom roles & maker-checker", YES, YES),
    ("feat",   "Advanced reports & Loyalty Pro", YES, YES),
    ("feat",   "API read access + daily auto-backup", YES, YES),
    ("section","SIMPLE BOOKKEEPING", "", ""),
    ("feat",   "Record income & expenses", NO, YES),
    ("feat",   "See money owed from charge sales", NO, YES),
    ("feat",   "Simple income-vs-expense summary", NO, YES),
    ("feat",   "Cash & e-wallet settlement view", NO, YES),
    ("section","FULL ACCOUNTING", "", ""),
    ("feat",   "Journal, chart of accounts, statements, BIR, AR/AP, period close", "Upgrade", "Upgrade"),
]

r = 2
for kind, a, b, c in rows:
    ca, cb, cc = ws.cell(r, 1, a), ws.cell(r, 2, b), ws.cell(r, 3, c)
    for cell in (ca, cb, cc):
        cell.border = border
    ca.alignment = left
    cb.alignment = center; cc.alignment = center
    if kind == "header":
        for cell in (ca, cb, cc):
            cell.font = f(12, True, BROWN_DK); cell.fill = PatternFill("solid", fgColor=CREAM)
        ws.row_dimensions[r].height = 24
    elif kind == "price":
        ca.font = f(11, True)
        for cell in (cb, cc):
            cell.font = f(14, True, BROWN); cell.number_format = '"₱"#,##0'
        ws.row_dimensions[r].height = 26
    elif kind == "text":
        ca.font = f(10, True); cb.font = f(10); cc.font = f(10)
    elif kind == "section":
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
        ca.value = a; ca.font = f(10, True, WHITE)
        ca.fill = PatternFill("solid", fgColor=BROWN); ca.alignment = left
    else:  # feat
        ca.font = f(10)
        for cell in (cb, cc):
            txt = str(cell.value)
            if txt == YES: cell.font = f(11, True, GREEN)
            elif txt == NO: cell.font = f(11, color=GREY)
            else: cell.font = f(9, True, BROWN_DK)
        if r % 2 == 1:
            for cell in (ca, cb, cc):
                if cell.fill.fgColor.rgb in (None, "00000000"):
                    cell.fill = PatternFill("solid", fgColor=CREAM_LT)
    r += 1

note_r = r + 1
ws.merge_cells(start_row=note_r, start_column=1, end_row=note_r, end_column=3)
ws.cell(note_r, 1, "All prices in Philippine pesos, per month, exclusive of 12% VAT. Solo Books upgrades to full accounting anytime — your data carries over.")
ws.cell(note_r, 1).font = f(8, color="6B7280"); ws.cell(note_r, 1).alignment = left

# ─────────────────────────── Sheet 2: All plans (reference) ──────────────────
ws2 = wb.create_sheet("All plans (reference)")
ws2.sheet_view.showGridLines = False
widths = [16, 22, 22, 22, 11, 11, 14, 16]
for i, w in enumerate(widths, 1):
    ws2.column_dimensions[chr(64 + i)].width = w

headers = ["Plan code", "Name", "Status", "Modules", "Base seats", "Max seats", "Monthly (₱)", "Annual prepay (₱)"]
for i, h in enumerate(headers, 1):
    c = ws2.cell(1, i, h)
    c.font = f(10, True, WHITE); c.fill = PatternFill("solid", fgColor=BROWN)
    c.alignment = center; c.border = border
ws2.row_dimensions[1].height = 30

def status(code):
    if code in ("SOLO_PRO", "SOLO_BOOKS"): return "Active"
    if code in ("SOLO_LITE", "SOLO_STANDARD"): return "Legacy (grandfathered)"
    return "Parked (legacy)"

def modules(code, c):
    if c["moduleCount"] == 3: return "All 3 modules"
    if c["moduleCount"] == 2: return "Any 2 modules"
    return "POS + simple ledger" if code == "SOLO_BOOKS" else "POS only"

# Active first, then legacy, then parked — matches business priority.
order = ["SOLO_PRO", "SOLO_BOOKS", "SOLO_LITE", "SOLO_STANDARD",
         "PAIR_T1", "PAIR_T2", "PAIR_T3", "SUITE_T1", "SUITE_T2", "SUITE_T3", "ENTERPRISE"]

r = 2
for code in order:
    c = caps[code]
    is_ent = code == "ENTERPRISE"
    vals = [code, c["label"], status(code), modules(code, c), c["baseSeats"], c["maxTotal"]]
    for i, v in enumerate(vals, 1):
        cell = ws2.cell(r, i, v)
        cell.font = f(10, b=(i == 1)); cell.border = border
        cell.alignment = left if i in (1, 2, 3, 4) else center
    # Monthly
    mcell = ws2.cell(r, 7)
    mcell.border = border; mcell.alignment = center
    if is_ent:
        mcell.value = "Custom"; mcell.font = f(10)
    else:
        mcell.value = c["monthly"]; mcell.font = f(10); mcell.number_format = '"₱"#,##0'
    # Annual prepay = monthly × annualMonths (computed value — static reference
    # sheet, no live model; avoids un-recalced formulas in non-Excel viewers).
    acell = ws2.cell(r, 8)
    acell.border = border; acell.alignment = center
    if is_ent:
        acell.value = "Custom"; acell.font = f(10)
    else:
        acell.value = c["monthly"] * c["annualMonths"]; acell.font = f(10); acell.number_format = '"₱"#,##0'
    r += 1

ws2.cell(r + 1, 1, "Annual prepay = monthly × 10 (2 months free). Source: PLAN_CAPS, packages/shared-types/src/plans.ts.")
ws2.cell(r + 1, 1).font = f(8, color="6B7280")

out = os.path.join(ROOT, "docs", "Clerque-Pricing-Tiers.xlsx")
wb.save(out)
print("Wrote", out)
